# -*- coding: utf-8 -*-
"""Export the Kitchen and Bathroom Nested Family Audit Notion DBs to .xlsx."""
import json, urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

TOKEN = open(r"N:\Design Technology Resources\01_BIM CONTENT\Content Conformance\notion_token.txt").read().strip()
API = "https://api.notion.com/v1"; V = "2022-06-28"
OUT = r"N:\Design Technology Resources\01_BIM CONTENT\Content Conformance\Document"
DBS = [
    ("Kitchen Nested Family Audit", "3a3d917d-ca75-81f7-8e23-f779ce6fe15f"),
    ("Bathroom Nested Family Audit", "3a3d917d-ca75-81ce-b86b-d06386b2e588"),
]
# display order of columns
COLS = ["Family Name", "Proposed Name", "Status", "Revit Category",
        "File Size (KB)", "Types", "Parent Family", "Child Families", "Child Count", "File Location"]


def req(m, u, p=None):
    r = urllib.request.Request(u, data=json.dumps(p).encode() if p else None, method=m,
                               headers={"Authorization": "Bearer %s" % TOKEN, "Notion-Version": V,
                                        "Content-Type": "application/json"})
    return json.loads(urllib.request.urlopen(r).read().decode())


def cell(prop):
    t = prop.get("type")
    if t == "title":
        return "".join(x["plain_text"] for x in prop["title"])
    if t == "rich_text":
        return "".join(x["plain_text"] for x in prop["rich_text"])
    if t == "multi_select":
        return ", ".join(o["name"] for o in prop["multi_select"])
    if t == "select":
        return prop["select"]["name"] if prop["select"] else ""
    if t == "number":
        return prop["number"]
    return ""


for name, db in DBS:
    rows, cur = [], None
    while True:
        pl = {"page_size": 100}
        if cur:
            pl["start_cursor"] = cur
        d = req("POST", "%s/databases/%s/query" % (API, db), pl)
        rows += d["results"]
        if not d.get("has_more"):
            break
        cur = d["next_cursor"]

    data = []
    for p in rows:
        props = p["properties"]
        data.append([cell(props.get(c, {})) for c in COLS])
    data.sort(key=lambda r: str(r[0]).lower())  # by Family Name

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]
    ws.append(COLS)
    hfill = PatternFill("solid", fgColor="1F3864")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(vertical="center")
    for row in data:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s1" % chr(ord("A") + len(COLS) - 1)
    # column widths
    widths = [34, 34, 16, 18, 12, 40, 46, 46, 12, 70]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    path = "%s\\%s.xlsx" % (OUT, name)
    wb.save(path)
    print("wrote %s  (%d rows)" % (path, len(data)))
